import openpyxl.utils.exceptions
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side, PatternFill
from datetime import datetime

title_style = NamedStyle(name="title_style")
title_style.font = Font(name="Arial")
title_style.border = Border(left=Side(style="thick"), right=Side(style="thick"),
                            top=Side(style="thick"), bottom=Side(style="thick"))
title_style.fill = PatternFill(fill_type='solid', start_color='b4b4b4')
title_style.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
# --------------------------------------------------
normal_style = NamedStyle(name="normal_style")
normal_style.number_format = 'mm-dd-yy'
normal_style.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
normal_style.fill = PatternFill(fill_type='solid', start_color='f0f0f0')
normal_style.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)


def style(new_rows):
    min_rows = ws.max_row - new_rows + 1
    for row in ws.iter_rows(min_row=min_rows, max_row=ws.max_row, min_col=2):
        for cell in row:
            cell.style = "normal_style"

    for row in ws["B2:D2"]:
        for cell in row:
            try:
                cell.style = title_style
            except ValueError:
                pass

    for row in range(min_rows, ws.max_row + 1):
        ws.row_dimensions[row].height = 14.40


if __name__ == '__main__':
    print("Program currently only work when notes have format:\n"
          "'Day' Date) Title\n"
          "(body)\n"
          "'Day' Date Title\n"
          "(body)\n")

    excel_name = input("Type name of the Excel you want to open or create: ")
    while True:
        try:
            new_wb = load_workbook(filename=excel_name)
        except openpyxl.utils.exceptions.InvalidFileException:
            print("Adding .xlsx at the end")
            excel_name = excel_name + ".xlsx"
        except FileNotFoundError:
            # if Excel file doesn't exist
            print("Creating new .xlsx file")
            new_wb = Workbook()
            new_ws = new_wb.active
            new_ws.title = "Journey"
            new_ws["B2"] = "Date"
            new_ws["C2"] = "Title"
            new_ws["D2"] = "Journal"
            new_wb.add_named_style(title_style)
            new_wb.add_named_style(normal_style)
            new_ws.row_dimensions[2].height = 20
            new_ws.column_dimensions['B'].width = 12
            new_ws.column_dimensions['C'].width = 25
            new_ws.column_dimensions['D'].width = 40
            new_wb.save(excel_name)
            print("New excel file saved")
            break
        else:
            break

    wb = load_workbook(filename=excel_name)
    ws = wb["Journey"]

    # Opening text file
    new_rows_amount = 0
    journal_date, journal_title, journal_body = "", "", ""
    # notes_file_path = input("Type path to your notes: ")
    notes_file_path = "C:/Users/okr65/Desktop/Journey10.txt"
    try:
        file = open(notes_file_path, "r", encoding="utf-8")
        for line in file:
            if line.lower().startswith("day"):
                if journal_body != "":
                    journal_body = journal_body.strip("\n")
                    data = [None, journal_date, journal_title, journal_body]
                    ws.append(data)
                    new_rows_amount += 1
                    journal_body = ""
                line = line.removeprefix("Day").strip(" ")
                journal_date = datetime.strptime(line[0:10], '%d.%m.%Y').date()
                journal_title = line[11:len(line)]
            else:
                journal_body += line
        style(new_rows_amount)
        file.close()
        wb.save(excel_name)
    except PermissionError:
        print("Can't modify Excel file while is open!")
