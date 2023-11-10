from openpyxl.styles import Font, Alignment, NamedStyle, Border, Side, PatternFill

title_style = NamedStyle(name="title_style")
title_style.font = Font(name="Arial")
title_style.border = Border(left=Side(style="thick"), right=Side(style="thick"),
                            top=Side(style="thick"), bottom=Side(style="thick"))
title_style.fill = PatternFill(fill_type='solid', start_color='b4b4b4')
title_style.alignment = Alignment(horizontal='center', vertical='center',
                                  wrap_text=True)
# --------------------------------------------------
normal_style = NamedStyle(name="normal_style")
normal_style.number_format = 'mm-dd-yy'
normal_style.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))
normal_style.fill = PatternFill(fill_type='solid', start_color='f0f0f0')
normal_style.alignment = Alignment(horizontal='center', vertical='center',
                                   wrap_text=True)


def style(new_rows, ws, wb):
    try:
        wb.add_named_style(title_style)
        wb.add_named_style(normal_style)
    except ValueError:
        # Styles already exists in Excel
        # There are moments when Excel diary doesn't exist but named styles exists.
        pass

    min_rows = ws.max_row - new_rows + 1
    for row in ws.iter_rows(min_row=min_rows, max_row=ws.max_row, min_col=2):
        for cell in row:
            try:
                cell.style = "normal_style"
            except ValueError:
                cell.style = normal_style

    for row in ws["B2:D2"]:
        for cell in row:
            try:
                cell.style = "title_style"
            except ValueError:
                cell.style = title_style

    for row in range(min_rows, ws.max_row + 1):
        ws.row_dimensions[row].height = 14.40
